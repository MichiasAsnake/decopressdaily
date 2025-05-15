import sys
import subprocess
import tempfile
import tkinter as tk
from tkinter import simpledialog
from playwright.sync_api import sync_playwright
import pandas as pd
import os
import shutil
import re
from datetime import datetime
from openpyxl import load_workbook
from utils import (
    get_login_info, get_clean_text, get_download_path, 
    get_current_date_formatted, LOGIN_URL, DASHBOARD_URL
)

def ensure_browser_installed():
    """Ensure we can use a browser in bundled app"""
    if getattr(sys, 'frozen', False):
        print("Running in a bundled application - checking browser setup...")
        
        # Set a temp directory for Playwright browsers if needed
        temp_dir = tempfile.mkdtemp(prefix="decopress_browser_")
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = temp_dir
        
        # Look for Chrome/Edge installations on the system
        chrome_paths = [
            # Regular Chrome paths
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            # Microsoft Edge paths
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        ]
        
        # Check which browsers are available
        available_browsers = []
        for path in chrome_paths:
            if os.path.exists(path):
                available_browsers.append(path)
                print(f"Found browser: {path}")
        
        if available_browsers:
            os.environ["PLAYWRIGHT_CHROME_EXECUTABLE_PATH"] = available_browsers[0]
            print(f"Using browser: {available_browsers[0]}")
            return available_browsers[0]
        else:
            print("No system browsers found - will try using Playwright's default approach")
            return None
    else:
        print("Running in development mode, using normal Playwright setup")
        return None

def ensure_paged_mode(page):
    """Ensure the page is in paged mode, not infinite scroll"""
    print("Checking if we need to enable paged mode...")
    
    try:
        # Find and click the settings button
        settings_button = page.query_selector('a[data-event="cw:list-settings"]')
        if not settings_button:
            print("Could not find settings button, will assume paged mode is active")
            return
            
        # Check if settings is already expanded
        is_expanded = settings_button.get_attribute('aria-expanded') == 'true'
        if not is_expanded:
            print("Clicking settings button to open settings panel")
            settings_button.click()
            # Give time for the panel to open
            page.wait_for_timeout(1000)
            
        # Look for both radio buttons
        paged_radio = page.query_selector('input[name="list-mode"][value="PAGED"]')
        infinite_radio = page.query_selector('input[name="list-mode"][value="INFINITE"]')
        
        if not paged_radio or not infinite_radio:
            print("Could not find list mode radio buttons, will continue assuming paged mode")
            # Try to close the panel if we opened it
            if not is_expanded:
                close_button = page.query_selector('.js-close-popup')
                if close_button:
                    close_button.click()
            return
        
        # Check if paged mode is already checked
        paged_checked = paged_radio.get_attribute('checked') == 'checked'
        
        if paged_checked:
            print("Paged mode is already active")
            # Close the panel if we opened it
            if not is_expanded:
                close_button = page.query_selector('.js-close-popup')
                if close_button:
                    close_button.click()
            return
            
        # If infinite mode is active, we need to click the paged mode label
        print("Infinite mode is active, switching to paged mode")
        paged_label = page.query_selector('label:has(input[name="list-mode"][value="PAGED"])')
        if paged_label:
            paged_label.click()
            print("Clicked paged mode option")
            
            # Wait for page to reload after changing the setting
            page.wait_for_load_state('networkidle')
            page.wait_for_timeout(2000)
            print("Page should now be in paged mode")
            return
            
        print("Could not find paged mode label to click, will continue with current mode")
    except Exception as e:
        print(f"Error ensuring paged mode: {str(e)}")
        # Continue with the current mode rather than failing
    finally:
        # Try to close the panel if it might be open
        try:
            close_button = page.query_selector('.js-close-popup')
            if close_button:
                close_button.click()
                page.wait_for_timeout(500)
        except:
            pass

def apply_patch_supply_filter(page):
    """Apply the PATCH SUPPLY -PS - GAMMA filter before scraping orders"""
    print("Applying PATCH SUPPLY -PS - GAMMA filter...")
    
    try:
        # Wait to make sure favorites are loaded
        page.wait_for_load_state('networkidle')
        page.wait_for_timeout(1000)
        
        # Multiple approaches to find and click the filter
        methods = [
            # Method 1: Find by data-id
            lambda: page.query_selector('label:has(input[data-id="6699e45c-7880-4fb2-9c60-ac8a6ad19de1"])'),
            # Method 2: Find by text content
            lambda: page.query_selector('label:text("PATCH SUPPLY -PS - GAMMA")'),
            # Method 3: Find by data-label attribute
            lambda: page.query_selector('label[data-label="PATCH SUPPLY -PS - GAMMA"]'),
            # Method 4: Find by text containing the phrase
            lambda: page.query_selector('label:has-text("PATCH SUPPLY -PS - GAMMA")'),
        ]
        
        # Try each method
        for method in methods:
            try:
                element = method()
                if element:
                    print(f"Found filter element using method {methods.index(method) + 1}")
                    element.click()
                    page.wait_for_load_state('networkidle')
                    page.wait_for_timeout(2000)  # Wait for filter to apply
                    
                    # Verify filter was applied - look for visible indication
                    active_filters = page.query_selector_all('.active-filter')
                    if active_filters:
                        print(f"Filter appears to be applied successfully ({len(active_filters)} active filters)")
                        return True
                    else:
                        print("Filter may not have been applied (no active filters detected)")
            except Exception as e:
                print(f"Method {methods.index(method) + 1} failed: {str(e)}")
                continue
        
        # If all methods failed, try direct JavaScript approach
        try:
            # Try to click the filter using JavaScript
            print("Attempting to apply filter via JavaScript...")
            js_success = page.evaluate('''() => {
                const elements = Array.from(document.querySelectorAll('label'));
                const filterLabel = elements.find(el => 
                    el.textContent.includes('PATCH SUPPLY -PS - GAMMA') || 
                    el.getAttribute('data-label') === 'PATCH SUPPLY -PS - GAMMA'
                );
                if (filterLabel) {
                    filterLabel.click();
                    return true;
                }
                return false;
            }''')
            
            if js_success:
                print("Filter applied via JavaScript")
                page.wait_for_load_state('networkidle')
                page.wait_for_timeout(2000)
                return True
        except Exception as e:
            print(f"JavaScript approach failed: {str(e)}")
        
        print("⚠️ Could not find or apply PATCH SUPPLY filter")
        return False
    except Exception as e:
        print(f"❌ Error applying filter: {str(e)}")
        return False

def extract_process_codes(row):
    """Extract process codes and quantities from a single row"""
    process_codes = []
    highest_qty = 0  # Track highest quantity for this row
    
    try:
        # Critical: Only look at this specific row - not any other rows on the page
        # This should limit our search to the current job only
        
        # First method: Use JavaScript but scope it specifically to this row
        result = row.evaluate("""(row) => {
            const processData = { codes: [], highestQty: 0 };
            
            // Only look for badges within THIS row, not on the entire page
            const badgeContainers = row.querySelectorAll('.ew-badge-container.process-codes, .process-codes');
            console.log('Found badge containers in row:', badgeContainers.length);
            
            for (const container of badgeContainers) {
                const badges = container.querySelectorAll('.ew-badge');
                console.log('Found badges in container:', badges.length);
                
                for (const badge of badges) {
                    const codeElement = badge.querySelector('.process-code-badge');
                    const qtyElement = badge.querySelector('.process-qty');
                    
                    if (codeElement) {
                        const code = codeElement.textContent.trim();
                        processData.codes.push(code);
                        console.log('Found code:', code);
                        
                        // Get quantity if available
                        if (qtyElement) {
                            const qtyText = qtyElement.textContent.trim();
                            const qty = parseInt(qtyText);
                            console.log('Found qty for code', code, ':', qty);
                            if (!isNaN(qty) && qty > processData.highestQty) {
                                processData.highestQty = qty;
                            }
                        }
                    }
                }
            }
            return processData;
        }""", row)  # Pass the row as an argument to the JavaScript function
        
        if result:
            process_codes = result['codes']
            highest_qty = result['highestQty']
            print(f"Row extraction result - Codes: {process_codes}, Highest Qty: {highest_qty}")
        
        # If JavaScript approach didn't work, fall back to a more direct approach
        if not process_codes or highest_qty == 0:
            # Try a direct approach using Playwright's API
            # Get process codes directly
            code_badges = row.query_selector_all(".process-code-badge")
            for badge in code_badges:
                code = badge.inner_text().strip()
                if code:
                    process_codes.append(code)
            
            # Try to find quantities directly in the row
            qty_elements = row.query_selector_all(".process-qty")
            for qty_element in qty_elements:
                qty_text = qty_element.inner_text().strip()
                try:
                    qty = int(qty_text)
                    if qty > highest_qty:
                        highest_qty = qty
                except ValueError:
                    continue
            
            print(f"Direct extraction - Codes: {process_codes}, Highest Qty: {highest_qty}")
        
    except Exception as e:
        print(f"Error extracting process codes and quantities from row: {str(e)}")
    
    return process_codes, highest_qty

def check_hw_garment_details(page, job_number):
    """
    Check garment details for HW jobs by clicking into the job page
    Returns the appropriate letter code based on garment material
    """
    try:
        # Construct the job URL
        job_url = f"https://intranet.decopress.com/Jobs/job.aspx?ID={job_number}"
        print(f"Navigating to job page: {job_url}")
        
        # Store the current URL to go back later
        current_url = page.url
        
        # Navigate to the job page
        page.goto(job_url)
        page.wait_for_load_state('networkidle')
        page.wait_for_selector("table", state="visible", timeout=10000)
        
        # Look for the jobline rows
        jobline_rows = page.query_selector_all("tr.js-jobline-row")
        
        # Keywords for ETCH
        etch_keywords = ['FAUX', 'LEATHER', 'LEATHERETTE', 'SUEDE', 'DENIM']
        
        # Keywords for SUB
        sub_keywords = ['SIMWOVEN', 'WOVEN', 'DECO TWILL', 'DECOTWILL', 'TWILL']
        
        # Keywords for EMB - highest priority
        emb_keywords = ['EMB', 'EMBROIDERY', 'EMBROIDERED']
        
        # Track what we find in all rows
        found_emb = False
        found_etch = False
        found_sub = False
        
        # Check each row for garment details
        for row in jobline_rows:
            try:
                # Try to get the data-garment attribute
                garment_attr = row.get_attribute('data-garment')
                if garment_attr:
                    garment_text = garment_attr.upper()
                    print(f"Found garment text: {garment_text}")
                    
                    # Check for EMB in this row
                    row_has_emb = any(keyword in garment_text for keyword in emb_keywords)
                    if row_has_emb:
                        found_emb = True
                        print(f"Row has EMB in garment: {garment_text}")
                    
                    # Check for ETCH keywords in this row
                    row_has_etch = any(keyword in garment_text for keyword in etch_keywords)
                    if row_has_etch:
                        found_etch = True
                        print(f"Row has ETCH material: {garment_text}")
                    
                    # Check for SUB keywords in this row
                    row_has_sub = any(keyword in garment_text for keyword in sub_keywords)
                    if row_has_sub:
                        found_sub = True
                        print(f"Row has SUB material: {garment_text}")
            except Exception as e:
                print(f"Error checking jobline row: {str(e)}")
        
        # Check garment cells directly if data-attribute approach didn't find everything
        if not (found_emb or found_etch or found_sub):
            garment_cells = page.query_selector_all("td.jobline-garment")  # More specific selector
            for cell in garment_cells:
                try:
                    cell_text = cell.inner_text().strip().upper()
                    
                    # Check for EMB in this cell
                    if any(keyword in cell_text for keyword in emb_keywords):
                        found_emb = True
                        print(f"Cell has EMB: {cell_text}")
                    
                    # Check for ETCH keywords in this cell
                    if any(keyword in cell_text for keyword in etch_keywords):
                        found_etch = True
                        print(f"Cell has ETCH material: {cell_text}")
                    
                    # Check for SUB keywords in this cell
                    if any(keyword in cell_text for keyword in sub_keywords):
                        found_sub = True
                        print(f"Cell has SUB material: {cell_text}")
                except Exception as e:
                    print(f"Error checking garment cell: {str(e)}")
        
        # Go back to the previous page
        page.goto(current_url)
        page.wait_for_load_state('networkidle')
        
        # Determine the final letter code based on what was found across all rows
        if found_emb and found_etch:
            return "EMB/ETCH"
        elif found_emb:
            return "EMB"
        elif found_etch:
            return "ETCH"
        elif found_sub:
            return "SUB"
        else:
            # Default to SUB if no specific material is found
            return "SUB"
    except Exception as e:
        print(f"Error checking HW garment details: {str(e)}")
        
        # Try to go back to the previous page
        try:
            page.goto(current_url)
            page.wait_for_load_state('networkidle')
        except:
            print("Error returning to previous page")
            
        # Default to SUB if there was an error
        return "SUB"

def determine_letter_code(page, process_codes, description, job_number):
    """Determine the letter code based on process codes and description"""
    # Convert to uppercase for case-insensitive comparison
    description_upper = description.upper() if description else ""
    process_codes_upper = [code.upper() for code in process_codes]
    
    # Flag to track if we need to check HW details
    has_hw = "HW" in process_codes_upper
    
    # Check immediate classifications first, but mark HW for later inspection
    if "AP" in process_codes_upper and "EM" in process_codes_upper:
        if has_hw:
            # Mark for HW inspection even though it has EM
            return "HW/EMB"
        return "SUB/EMB"
    elif "AP" in process_codes_upper:
        if has_hw:
            # Mark for HW inspection even though it has AP
            return "HW/SUB"
        return "SUB"
    elif "EM" in process_codes_upper:
        if has_hw:
            # Mark for HW inspection even though it has EM
            return "HW/EMB"
        return "EMB"
    elif "DS" in process_codes_upper:
        if has_hw:
            # Mark for HW inspection even though it has DS
            return "HW/ETCH"
        return "ETCH"
    # Check for HW as last priority if no other codes matched
    elif has_hw:
        # Initially just mark as HW, we'll check details later
        return "HW"
    
    # Default if nothing matches
    return ""

def has_paplique(process_codes):
    """Check if the process Has Patch Apply"""
    process_codes_upper = [code.upper() for code in process_codes]
    return "PA" in process_codes_upper

def get_short_description(full_description):
    """Get the first four words from a description"""
    if not full_description:
        return ""
        
    # Split by spaces
    words = full_description.split()
    
    # Take up to first 4 words
    short_words = words[:4]
    
    # Join back with spaces
    return " ".join(short_words)

def extract_location_tags(row):
    """Extract location tags from jobtag-container"""
    location = ""
    
    try:
        # Use JavaScript to extract tags from the jobtag-container
        result = row.evaluate("""(row) => {
            // Look for the jobtag-container within this row
            const tagContainer = row.querySelector('.jobtag-container');
            if (!tagContainer) return null;
            
            // Find all tag elements
            const tagElements = tagContainer.querySelectorAll('li .jobtag.tag.showtag .tag-text');
            const tags = Array.from(tagElements).map(el => el.textContent.trim().toLowerCase());
            
            // Return the array of tags
            return tags;
        }""", row)
        
        if result:
            # Define priority order - earlier items take precedence
            priority_tags = ["rfp", "@sub", "@laser", "qc"]
            priority_codes = {"rfp": "RFP", "@sub": "SUB", "@laser": "LASER", "qc": "QC"}
            
            # Find the highest priority tag
            for tag in priority_tags:
                if tag in result:
                    location = priority_codes[tag]
                    print(f"Found location tag: {tag} -> {location}")
                    break
            
            print(f"All job tags found: {result}, Selected location: {location}")
    
    except Exception as e:
        print(f"Error extracting location tags: {str(e)}")
    
    return location

def scrape_orders(page):
    orders = []
    current_page = 1
    max_pages = 3  # Increased to 3 pages
    max_orders = 31  # Increased to 31 orders
    visited_pages = 0  # Track actual pages visited
    
    # Wait for the table to be present and visible
    page.wait_for_selector("table.data-results", state="visible", timeout=30000)
    print("Table found, starting to scrape...")
    
    while visited_pages < max_pages and len(orders) < max_orders:
        print(f"Processing page {current_page} (visited {visited_pages + 1} of {max_pages})")
        # Wait for any loading indicators to disappear
        page.wait_for_load_state('networkidle')
        
        rows = page.query_selector_all("table.data-results tbody tr")
        print(f"Found {len(rows)} rows on current page")
        
        for row in rows:
            if len(orders) >= max_orders:
                print(f"Reached maximum of {max_orders} orders")
                break
                
            try:
                days_element = row.query_selector("span.js-days-to-due-date")
                if not days_element:
                    print("Days element not found in row")
                    continue
                    
                days_text = days_element.inner_text().strip()
                print(f"Days text found: {days_text}")
                
                try:
                    days = int(days_text)
                except ValueError:
                    print(f"Could not convert days text to integer: {days_text}")
                    continue
                
                # Changed to include days 0-4
                if days in (0, 1, 2, 3, 4):
                    job_number = get_clean_text(row.query_selector("td:nth-child(1)"))
                    # Only keep numeric job numbers
                    if not job_number.isdigit():
                        continue
                        
                    description_element = row.query_selector("td:nth-child(3)")
                    full_description = get_clean_text(description_element)
                    
                    # Get first four words
                    short_description = get_short_description(full_description)
                    
                    # Extract process codes and their highest quantity
                    process_codes, highest_qty = extract_process_codes(row)
                    print(f"Job {job_number} - Process codes: {process_codes}, Highest quantity: {highest_qty}")
                    
                    # Extract location tags (rfp, @sub, @laser, qc)
                    location = extract_location_tags(row)
                    
                    # Determine letter code
                    letter_code = determine_letter_code(page, process_codes, full_description, job_number)
                    
                    # Check for applique
                    has_pa = has_paplique(process_codes)
                    
                    job_status = get_clean_text(row.query_selector("td:nth-child(4)"))
                    # Keep only text after hyphen if it exists
                    if " - " in job_status:
                        job_status = job_status.split(" - ")[1]
                    
                    order = {
                        "Job Number": job_number,
                        "Customer": get_clean_text(row.query_selector("td:nth-child(2)")),
                        "Description": full_description,
                        "Short Description": short_description,
                        "Job Status": job_status,
                        "Order #": get_clean_text(row.query_selector("td:nth-child(5)")),
                        "Date In": get_clean_text(row.query_selector("td:nth-child(6)")),
                        "Ship Date": get_clean_text(row.query_selector("td:nth-child(7)")),
                        "Days Remaining": days,
                        "Process Codes": process_codes,
                        "Letter Code": letter_code,
                        "Has Patch Apply": has_pa,
                        "Quantity": highest_qty,
                        "Location": location
                    }
                    orders.append(order)
                    print(f"Added order with {days} days remaining, Letter Code: {letter_code}, Has Patch Apply: {has_pa}")
            except Exception as e:
                print(f"Error processing row: {str(e)}")
                continue
        
        # Increment visited pages counter
        visited_pages += 1
        
        # Move to next page if we haven't reached max pages and max orders
        if visited_pages < max_pages and len(orders) < max_orders:
            try:
                # Find the next page link
                next_page = page.query_selector(f"ul.pagination li[data-lp='{current_page + 1}'] a.page-link")
                if next_page:
                    print(f"Clicking page {current_page + 1}")
                    next_page.click()
                    page.wait_for_timeout(2000)  # Wait for page load
                    current_page += 1
                else:
                    print("Next page link not found")
                    break
            except Exception as e:
                print(f"Error navigating to next page: {str(e)}")
                break
        elif len(orders) >= max_orders:
            print(f"Reached maximum of {max_orders} orders, stopping pagination")
        else:
            print("Reached maximum page limit")
            break
            
    print(f"Total orders found: {len(orders)}")
    
    # Now process any HW jobs to determine their actual letter code
    print("Processing HW jobs to determine material types...")
    for order in orders:
        # Check any order that has HW in its letter code
        if "HW" in order["Letter Code"]:
            process_codes = order["Process Codes"]
            job_number = order["Job Number"]
            description = order["Description"]
            
            # Get the original letter code
            original_code = order["Letter Code"]
            
            # Check the HW garment details
            hw_material_code = check_hw_garment_details(page, job_number)
            
            # Special handling for combined codes
            if original_code == "HW/EMB":
                # If the material is ETCH, combine EMB and ETCH
                if "ETCH" in hw_material_code:
                    order["Letter Code"] = "EMB/ETCH"
                else:
                    # Otherwise keep just EMB
                    order["Letter Code"] = "EMB"
            elif original_code == "HW/SUB":
                # If the material is ETCH, combine SUB and ETCH
                if "ETCH" in hw_material_code:
                    order["Letter Code"] = "SUB/ETCH"
                else:
                    # Otherwise keep just SUB
                    order["Letter Code"] = "SUB"
            elif original_code == "HW/ETCH":
                # Already has ETCH, just keep it
                order["Letter Code"] = "ETCH"
            else:
                # For plain HW, use the determined code
                order["Letter Code"] = hw_material_code
            
            print(f"Updated HW job {job_number} from {original_code} to {order['Letter Code']}")
    
    return orders

def create_daily_report(orders):
    """Create a daily report using the template"""
    if not orders:
        print("No orders to export")
        return None
    
    # Get the template path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, "DECOPRESS DAILY Template.xlsx")
    
    if not os.path.exists(template_path):
        print(f"❌ Template file not found: {template_path}")
        return None
    
    # Get the download path and create filename
    download_path = get_download_path()
    current_date = get_current_date_formatted()
    excel_filename = f"{current_date}_DECOPRESS_DAILY.xlsx"
    excel_filepath = os.path.join(download_path, excel_filename)
    
    try:
        # Copy the template to the new file
        shutil.copy2(template_path, excel_filepath)
        
        # Open the template with openpyxl
        workbook = load_workbook(excel_filepath)
        sheet = workbook.active
        
        # Get information about merged cells to avoid setting values to merged cells
        merged_cell_ranges = sheet.merged_cells.ranges
        merged_cells = []
        for merged_range in merged_cell_ranges:
            cells = []
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cells.append((row, col))
            merged_cells.extend(cells)
        print(f"Identified {len(merged_cells)} merged cells to avoid writing to")
        
        # Add today's date in a cell that's not merged
        # Try different locations for the date
        try:
            today = datetime.now()
            formatted_date = today.strftime("%m.%d.%y")  # Format as MM.DD.YY
            
            # Check if A3 is a merged cell
            is_merged_a3 = False
            for merged_range in sheet.merged_cells.ranges:
                if "A3" in merged_range.coord:
                    is_merged_a3 = True
                    break
            
            # Try different cells for the date
            date_cells = ["A3"]
            date_placed = False
            
            for cell_coord in date_cells:
                try:
                    # Check if this cell is part of a merged range
                    is_merged = False
                    for merged_range in sheet.merged_cells.ranges:
                        if cell_coord in merged_range.coord:
                            is_merged = True
                            break
                    
                    if not is_merged:
                        sheet[cell_coord] = f"Date: {formatted_date}"
                        date_placed = True
                        print(f"Date placed in {cell_coord}")
                        break
                except Exception as e:
                    print(f"Could not place date in {cell_coord}: {str(e)}")
            
            # If we couldn't place it in any of the preferred cells, try a different approach
            if not date_placed:
                # Find the main merge cell for D3 if it's merged
                if is_merged_a3:
                    for merged_range in sheet.merged_cells.ranges:
                        if "D3" in merged_range.coord:
                            main_cell = merged_range.coord.split(":")[0]  # Get the top-left cell of the merge
                            try:
                                sheet[main_cell] = f"Date: {formatted_date}"
                                date_placed = True
                                print(f"Date placed in merge cell {main_cell}")
                                break
                            except Exception as e:
                                print(f"Could not place date in main merge cell {main_cell}: {str(e)}")
            
            # Last resort - just note that we couldn't place the date
            if not date_placed:
                print("⚠️ Could not place date in any cell. Template may have unusual merged cells.")
        
        except Exception as e:
            print(f"Error adding date: {str(e)}")
            # Continue without the date rather than failing
        
        # Start row for data (B5 is where the first job number goes)
        current_row = 5
        
        # Map of column letters to their numerical indices
        col_map = {'B': 2, 'C': 3, 'D': 4, 'E': 5, 'F': 6, 'H': 8, 'I': 9}
        
        # Fill in the data for each order
        for order in orders:
            try:
                # Print all order fields for debugging
                print(f"Processing order: {order}")
                
                # Check if we're trying to write to a merged cell first
                if (current_row, col_map['B']) in merged_cells:
                    print(f"Row {current_row}, Col B is a merged cell - skipping")
                    current_row += 1
                    continue
                
                # Job Number in column B
                sheet.cell(row=current_row, column=col_map['B']).value = order.get("Job Number", "")
                
                # Short Description in column C
                sheet.cell(row=current_row, column=col_map['C']).value = order.get("Short Description", "")
                
                # Letter Code in column D
                sheet.cell(row=current_row, column=col_map['D']).value = order.get("Letter Code", "")
                
                # Location in column E
                sheet.cell(row=current_row, column=col_map['E']).value = order.get("Location", "")
                
                # Quantity in column F - make sure we're adding it correctly
                qty_value = order.get("Quantity", 0)
                print(f"Adding quantity for job {order.get('Job Number', '')}: {qty_value}")
                if qty_value and qty_value > 0:
                    sheet.cell(row=current_row, column=col_map['F']).value = qty_value
                
                # Has Patch Apply in column H (TRUE/FALSE)
                sheet.cell(row=current_row, column=col_map['H']).value = "TRUE" if order.get("Has Patch Apply") else "FALSE"
                
                # Days Remaining in column I
                sheet.cell(row=current_row, column=col_map['I']).value = order.get("Days Remaining", "")
                
                current_row += 1
            except Exception as e:
                print(f"Error adding row data for job {order.get('Job Number', '')}: {str(e)}")
                # Continue with next row rather than failing completely
        
        # Save the workbook
        workbook.save(excel_filepath)
        print(f"✅ Created daily report using template: {excel_filepath}")
        return excel_filepath
    except Exception as e:
        print(f"❌ Error creating report: {str(e)}")
        return None

def run():
    # Ensure browser is installed
    browser_path = ensure_browser_installed()
    report_path = None  # Initialize report path variable

    with sync_playwright() as p:
        # Launch browser using system installation if available
        launch_options = {
            "headless": False,
        }
        
        if browser_path:
            launch_options["executable_path"] = browser_path
            browser = p.chromium.launch(headless=True)
        else:
            # Let Playwright try to find its own browser
            browser = p.chromium.launch(headless=True)
            
        context = browser.new_context()
        page = context.new_page()

        try:
            # Login
            page.goto(LOGIN_URL)
            page.wait_for_load_state('networkidle')  # Wait for page to fully load
            
            username, password = get_login_info()
            
            # Check if login was cancelled
            if not username or not password:
                print("Login cancelled by user")
                browser.close()
                return None
                
            page.wait_for_selector("#txt_Username", timeout=60000)
            page.fill("#txt_Username", username)
            page.fill("#txt_Password", password)
            page.click("#btn_Login")
            page.wait_for_selector("#jobStatusListResults", timeout=10000)

            # Go to Job Status List
            page.goto(DASHBOARD_URL)
            page.wait_for_selector("table.data-results")
            
            # Ensure paged mode is active (not infinite scroll)
            ensure_paged_mode(page)

            # Apply the PATCH SUPPLY -PS - GAMMA filter
            filter_applied = apply_patch_supply_filter(page)
            if not filter_applied:
                print("⚠️ Continuing without filter")

            # Scrape
            print("Scraping urgent orders...")
            orders = scrape_orders(page)

            # Create report using template
            if orders:
                df = pd.DataFrame(orders)
                df.sort_values("Days Remaining", inplace=True)
                sorted_orders = df.to_dict('records')
                
                report_path = create_daily_report(sorted_orders)
                if report_path:
                    print(f"✅ Exported {len(orders)} urgent orders to Excel: {report_path}")
            else:
                print("⚠️ No 0, 1, 2, 3, or 4-day orders found.")

        except Exception as e:
            print(f"❌ Error: {str(e)}")
        finally:
            browser.close()
            
    return report_path

if __name__ == "__main__":
    run() 